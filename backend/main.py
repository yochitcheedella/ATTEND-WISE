from fastapi import FastAPI, Depends, HTTPException, status, File, UploadFile, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, time, timedelta, timezone, datetime
import os
import base64
import requests
import json
import math
import secrets
import hashlib
import firebase_admin
from firebase_admin import credentials, messaging
try:
    cred = credentials.Certificate('backend/firebase-adminsdk-key.json')
    firebase_admin.initialize_app(cred)
except Exception as e:
    print(f'Firebase Admin init failed (missing file or invalid key): {e}')

from fastapi.security import OAuth2PasswordRequestForm
from . import models, schemas, auth
from .database import engine, get_db

# Create database tables
models.Base.metadata.create_all(bind=engine)

# In-memory password reset token store: {token_hash: {email, expires_at}}
_reset_tokens: dict = {}

# Database migrations — add new columns safely (works with both SQLite and PostgreSQL)
from sqlalchemy import text, inspect

def _column_exists(conn, table_name: str, column_name: str, db_url: str) -> bool:
    """Check if a column exists — works for both SQLite and PostgreSQL."""
    if db_url.startswith("sqlite"):
        result = conn.execute(text(f"PRAGMA table_info({table_name})"))
        return any(row[1] == column_name for row in result.fetchall())
    else:
        result = conn.execute(text(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_name = :tbl AND column_name = :col"
        ), {"tbl": table_name, "col": column_name})
        return result.fetchone() is not None

try:
    from .database import SQLALCHEMY_DATABASE_URL as _db_url
    with engine.begin() as conn:
        # subjects table
        for col, coltype, default in [
            ("minimum_required_attendance", "FLOAT", "75.0"),
            ("subject_type", "TEXT", "'Theory'"),
            ("weekly_classes", "INTEGER", "4"),
            ("total_planned_classes", "INTEGER", "40"),
            ("baseline_conducted", "INTEGER", "0"),
            ("baseline_attended", "INTEGER", "0"),
            ("current_conducted", "INTEGER", "0"),
            ("current_attended", "INTEGER", "0"),
            ("last_synced_at", "TIMESTAMP", "NULL"),
        ]:
            if not _column_exists(conn, "subjects", col, _db_url):
                conn.execute(text(f"ALTER TABLE subjects ADD COLUMN {col} {coltype} DEFAULT {default}"))

        # attendance table
        if not _column_exists(conn, "attendance", "source", _db_url):
            conn.execute(text("ALTER TABLE attendance ADD COLUMN source TEXT DEFAULT 'daily_tracker'"))

        # semesters table
        if not _column_exists(conn, "semesters", "academic_calendar", _db_url):
            conn.execute(text("ALTER TABLE semesters ADD COLUMN academic_calendar TEXT"))

        # users table
        for col, coltype in [
            ("roll_number", "TEXT"),
            ("section", "TEXT"),
            ("year", "TEXT"),
            ("profile_photo", "TEXT"),
            ("register_number", "TEXT"),
            ("university", "TEXT"),
        ]:
            if not _column_exists(conn, "users", col, _db_url):
                conn.execute(text(f"ALTER TABLE users ADD COLUMN {col} {coltype}"))
except Exception as e:
    print("Migration warning:", e)

# Rate limiter setup
limiter = Limiter(key_func=get_remote_address)

app = FastAPI(
    title="AttendWise API",
    description="Backend API for AttendWise AI-Powered Student Attendance Companion",
    version="1.0.0"
)

app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Configure CORS
allowed_origins_str = os.getenv("ALLOWED_ORIGINS", "")
if allowed_origins_str:
    allowed_origins = [origin.strip() for origin in allowed_origins_str.split(",")]
else:
    # All valid origins: Vercel, Capacitor Android (all schemes), local dev, and Render
    allowed_origins = [
        "https://attend-wise.vercel.app",
        "https://attend-wise.onrender.com",
        # Android Capacitor WebView uses these schemes
        "capacitor://localhost",
        "https://localhost",
        "http://localhost",
        # Local dev ports
        "http://localhost:3000",
        "http://localhost:5173",
        "http://localhost:8000",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:8000",
    ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_origin_regex=r"https://.*\.vercel\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

import uuid
import time
from datetime import datetime
from fastapi.responses import JSONResponse
from fastapi.exceptions import RequestValidationError
from starlette.exceptions import HTTPException as StarletteHTTPException

# Middleware for Request ID & Structured Logging (Checklist Section 6)
@app.middleware("http")
async def add_request_id_and_logging(request, call_next):
    request_id = request.headers.get("X-Request-ID", str(uuid.uuid4()))
    request.state.request_id = request_id
    
    start_time = time.time()
    try:
        response = await call_next(request)
        process_time = time.time() - start_time
        print(f"INFO: {request.method} {request.url.path} - {response.status_code} (Duration: {process_time:.4f}s, RequestID: {request_id})")
        response.headers["X-Request-ID"] = request_id
        return response
    except Exception as e:
        process_time = time.time() - start_time
        print(f"ERROR: {request.method} {request.url.path} failed - {str(e)} (Duration: {process_time:.4f}s, RequestID: {request_id})")
        raise e

# Custom Global Exception Handlers (Checklist Section 5)
@app.exception_handler(StarletteHTTPException)
async def http_exception_handler(request, exc):
    return JSONResponse(
        status_code=exc.status_code,
        content={
            "success": False,
            "message": exc.detail,
            "error": exc.detail,
            "timestamp": datetime.utcnow().isoformat(),
            "requestId": getattr(request.state, "request_id", str(uuid.uuid4()))
        }
    )

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    errors = exc.errors()
    err_msgs = [f"{'.'.join(str(loc) for loc in err['loc'])}: {err['msg']}" for err in errors]
    message = "Validation error: " + "; ".join(err_msgs)
    return JSONResponse(
        status_code=422,
        content={
            "success": False,
            "message": message,
            "error": errors,
            "timestamp": datetime.utcnow().isoformat(),
            "requestId": getattr(request.state, "request_id", str(uuid.uuid4()))
        }
    )

@app.exception_handler(Exception)
async def generic_exception_handler(request, exc):
    import traceback
    traceback.print_exc()
    return JSONResponse(
        status_code=500,
        content={
            "success": False,
            "message": "An unexpected server error occurred.",
            "error": str(exc),
            "timestamp": datetime.utcnow().isoformat(),
            "requestId": getattr(request.state, "request_id", str(uuid.uuid4()))
        }
    )

from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

frontend_dist = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "dist")
if os.path.exists(frontend_dist):
    app.mount("/assets", StaticFiles(directory=os.path.join(frontend_dist, "assets")), name="assets")

# --- Health / Wake-up endpoints ---
@app.get("/ping", tags=["Health"])
def ping():
    """Ultra-fast endpoint to wake up the Render free-tier server."""
    return {"status": "ok", "message": "pong"}

@app.get("/health", tags=["Health"])
def health_check(db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT 1"))
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Database unhealthy: {e}")

# --- Authentication Endpoints ---
@app.post("/auth/register", response_model=schemas.User, tags=["Authentication"])
def register_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
    db_user = db.query(models.User).filter(models.User.email == user.email).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")

    hashed_password = auth.get_password_hash(user.password)
    new_user = models.User(
        name=user.name,
        email=user.email,
        password_hash=hashed_password,
        college=user.college,
        branch=user.branch,
        semester=user.semester,
        roll_number=user.roll_number,
        section=user.section,
        year=user.year,
        attendance_goal=user.attendance_goal
    )
    try:
        db.add(new_user)
        db.commit()
        db.refresh(new_user)
        return new_user
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create user profile: {str(e)}")

@app.post("/auth/login", response_model=schemas.Token, tags=["Authentication"])
@limiter.limit("10/minute")
def login_for_access_token(request: Request, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == form_data.username).first()
    if not user or not auth.verify_password(form_data.password, user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = auth.timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth.create_access_token(
        data={"sub": user.email}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

# --- Subjects Endpoints ---
@app.post("/subjects", response_model=schemas.Subject, tags=["Subjects"])
def create_subject(subject: schemas.SubjectCreate, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    db_subject = models.Subject(**subject.model_dump(), user_id=current_user.id)
    db.add(db_subject)
    db.commit()
    db.refresh(db_subject)
    return db_subject

@app.get("/subjects", response_model=List[schemas.Subject], tags=["Subjects"])
def get_subjects(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    return db.query(models.Subject).filter(models.Subject.user_id == current_user.id).all()

@app.put("/subjects/{subject_id}", response_model=schemas.Subject, tags=["Subjects"])
def update_subject(subject_id: int, subject_update: schemas.SubjectCreate, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    db_sub = db.query(models.Subject).filter(models.Subject.id == subject_id, models.Subject.user_id == current_user.id).first()
    if not db_sub:
        raise HTTPException(status_code=404, detail="Subject not found")
    
    # Update subject fields
    db_sub.name = subject_update.name
    db_sub.code = subject_update.code
    db_sub.prof = subject_update.prof
    db_sub.credits = subject_update.credits
    db_sub.color = subject_update.color
    db_sub.minimum_required_attendance = subject_update.minimum_required_attendance
    db_sub.subject_type = subject_update.subject_type
    db_sub.weekly_classes = subject_update.weekly_classes
    db_sub.total_planned_classes = subject_update.total_planned_classes
    
    db.commit()
    db.refresh(db_sub)
    return db_sub

@app.delete("/subjects/{subject_id}", tags=["Subjects"])
def delete_subject(subject_id: int, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    db_sub = db.query(models.Subject).filter(models.Subject.id == subject_id, models.Subject.user_id == current_user.id).first()
    if not db_sub:
        raise HTTPException(status_code=404, detail="Subject not found")
    
    # Check if there are associated timetables or attendances, we will delete them or cascade.
    # To keep database clean, delete related timetable and attendances for this subject.
    db.query(models.Timetable).filter(models.Timetable.subject_id == subject_id).delete()
    db.query(models.Attendance).filter(models.Attendance.subject_id == subject_id).delete()
    
    db.delete(db_sub)
    db.commit()
    return {"message": "Subject and all related data deleted successfully"}

@app.post("/subjects/{subject_id}/sync", response_model=schemas.Subject, tags=["Subjects"])
def sync_subject_attendance(subject_id: int, request: schemas.SyncAttendanceRequest, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    db_sub = db.query(models.Subject).filter(models.Subject.id == subject_id, models.Subject.user_id == current_user.id).first()
    if not db_sub:
        raise HTTPException(status_code=404, detail="Subject not found")
    
    try:
        db_sub.baseline_conducted = request.conducted
        db_sub.baseline_attended = request.attended
        db_sub.last_synced_at = datetime.now(timezone.utc)
        
        db.commit()
        db.refresh(db_sub)
        return db_sub
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

# --- Timetable Endpoints ---
@app.post("/timetable", response_model=schemas.Timetable, tags=["Timetable"])
def create_timetable(timetable: schemas.TimetableCreate, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    db_tt = models.Timetable(**timetable.model_dump(), user_id=current_user.id)
    db.add(db_tt)
    db.commit()
    db.refresh(db_tt)
    return db_tt

@app.get("/timetable", response_model=List[schemas.Timetable], tags=["Timetable"])
def get_timetable(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    return db.query(models.Timetable).filter(models.Timetable.user_id == current_user.id).all()

# --- Attendance Endpoints ---
@app.post("/attendance", response_model=schemas.Attendance, tags=["Attendance"])
def log_attendance(attendance: schemas.AttendanceCreate, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    db_att = models.Attendance(**attendance.model_dump(), user_id=current_user.id)
    db.add(db_att)
    db.commit()
    db.refresh(db_att)
    return db_att

@app.get("/attendance", response_model=List[schemas.Attendance], tags=["Attendance"])
def get_attendance(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    return db.query(models.Attendance).filter(models.Attendance.user_id == current_user.id).all()

# --- Helper Functions ---
def _compute_streak(user_id: int, db: Session) -> int:
    """
    Compute the current attendance streak:
    Count consecutive class days (going backwards from today) where the 
    user has at least one 'present' record.
    - Days with zero attendance records (weekends, holidays) are skipped.
    - Days where ALL records are 'cancelled' or 'holiday' are also skipped.
    - A day breaks the streak only if it has at least one 'absent' record
      AND zero 'present' records (i.e. the student was completely absent).
    - Today counts toward the streak if already partially marked present.
    """
    today = date.today()
    streak = 0
    check_date = today  # Start from today itself
    
    for _ in range(365):  # Max 1 year lookback
        day_attendances = db.query(models.Attendance).filter(
            models.Attendance.user_id == user_id,
            models.Attendance.date == check_date
        ).all()
        
        has_present = any(a.status.lower() in ("present", "late_entry", "od", "event_leave", "medical_leave") for a in day_attendances)
        has_absent = any(a.status.lower() == "absent" for a in day_attendances)
        
        # Check if it's a meaningful class day: has at least one present/absent record
        active_statuses = {a.status.lower() for a in day_attendances}
        is_class_day = bool(active_statuses - {"cancelled", "holiday", "upcoming"})
        
        if len(day_attendances) == 0 or not is_class_day:
            # Skip days with no records or only cancelled/holiday/upcoming
            check_date -= timedelta(days=1)
            continue
        elif has_present:
            # Student attended at least one class — counts as streak day
            streak += 1
            check_date -= timedelta(days=1)
        else:
            # Student was absent on a class day — streak broken
            break
    
    return streak


def _get_attendance_weight(att: models.Attendance, subject_map: dict = None) -> int:
    # If the subject is a lab/practical, weight = 3, else 1
    if subject_map and att.subject_id in subject_map:
        sub = subject_map[att.subject_id]
        if sub.subject_type.lower() == 'practical' or 'lab' in sub.name.lower():
            return 3
    return 1

def _compute_global_stats(user_id: int, db: Session):
    """Compute overall attendance stats across all subjects."""
    attendances = db.query(models.Attendance).filter(
        models.Attendance.user_id == user_id
    ).all()
    subjects = db.query(models.Subject).filter(models.Subject.user_id == user_id).all()
    sub_map = {s.id: s for s in subjects}
    
    present = sum(_get_attendance_weight(a, sub_map) for a in attendances if a.status.lower() in ("present", "late_entry", "od", "event_leave", "medical_leave"))
    absent = sum(_get_attendance_weight(a, sub_map) for a in attendances if a.status.lower() == "absent")
    total = present + absent
    percentage = round((present / total * 100), 2) if total > 0 else 0.0
    
    return {"present": present, "absent": absent, "total": total, "percentage": percentage}

def _compute_safe_bunks(present: int, total: int, target_pct: float) -> dict:
    """Compute safe bunks or classes needed to reach target."""
    if total == 0:
        return {"type": "neutral", "count": 0, "text": "No data", "desc": "Start marking attendance to view limits."}
    
    target_fraction = target_pct / 100.0
    safe_bunks = math.floor((present - target_fraction * total) / target_fraction)
    
    if safe_bunks >= 0:
        return {
            "type": "safe",
            "count": safe_bunks,
            "text": f"{safe_bunks} Classes",
            "desc": f"You can safely miss {safe_bunks} consecutive classes while staying above {target_pct}%."
        }
    else:
        classes_needed = math.ceil((target_fraction * total - present) / (1 - target_fraction))
        return {
            "type": "risk",
            "count": classes_needed,
            "text": f"{classes_needed} Classes Required",
            "desc": f"Warning! You must attend the next {classes_needed} consecutive classes to reach {target_pct}%."
        }

# --- Analytics Endpoints ---
@app.get("/analytics/dashboard", tags=["Analytics"])
def get_dashboard_analytics(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    stats = _compute_global_stats(current_user.id, db)
    bunk_info = _compute_safe_bunks(stats["present"], stats["total"], current_user.attendance_goal)
    return {
        "overall_percentage": stats["percentage"],
        "total_classes": stats["total"],
        "present": stats["present"],
        "absent": stats["absent"],
        "safe_bunks": bunk_info["count"] if bunk_info["type"] == "safe" else 0,
        "bunk_analysis": bunk_info
    }

@app.get("/analytics/detailed", tags=["Analytics"])
def get_detailed_analytics(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    """
    Returns:
    - Weekly trend: last 6 weeks attendance percentage
    - Heatmap data: attendance density per week for last 16 weeks
    - Subject-wise detailed stats
    """
    user_id = current_user.id
    today = date.today()
    
    # --- Weekly Trend (last 6 weeks) ---
    weekly_trend = []
    for week_offset in range(5, -1, -1):  # from 5 weeks ago to current week
        week_start = today - timedelta(days=today.weekday()) - timedelta(weeks=week_offset)
        week_end = week_start + timedelta(days=6)
        
        week_attendances = db.query(models.Attendance).filter(
            models.Attendance.user_id == user_id,
            models.Attendance.date >= week_start,
            models.Attendance.date <= week_end
        ).all()
        
        subjects = db.query(models.Subject).filter(models.Subject.user_id == user_id).all()
        sub_map = {s.id: s for s in subjects}
        p = sum(_get_attendance_weight(a, sub_map) for a in week_attendances if a.status.lower() in ("present", "late_entry", "od", "event_leave", "medical_leave"))
        total = sum(_get_attendance_weight(a, sub_map) for a in week_attendances if a.status.lower() in ("present", "late_entry", "od", "event_leave", "medical_leave", "absent"))
        pct = round(p / total * 100, 1) if total > 0 else 0.0
        weekly_trend.append(pct)
    
    # --- Heatmap Data (last 16 weeks, daily density) ---
    heatmap_start = today - timedelta(weeks=16)
    heatmap_attendances = db.query(models.Attendance).filter(
        models.Attendance.user_id == user_id,
        models.Attendance.date >= heatmap_start
    ).all()
    
    # Build a dict: date_str -> intensity (0-4)
    from collections import defaultdict
    date_stats = defaultdict(lambda: {"present": 0, "total": 0})
    for a in heatmap_attendances:
        d = a.date.isoformat()
        if a.status.lower() in ("present", "late_entry", "od", "event_leave", "medical_leave", "absent"):
            date_stats[d]["total"] += 1
        if a.status.lower() in ("present", "late_entry", "od", "event_leave", "medical_leave"):
            date_stats[d]["present"] += 1
    
    heatmap = {}
    for d, stats in date_stats.items():
        if stats["total"] == 0:
            heatmap[d] = 0
        else:
            pct = stats["present"] / stats["total"]
            if pct == 0:
                heatmap[d] = 0
            elif pct < 0.25:
                heatmap[d] = 1
            elif pct < 0.5:
                heatmap[d] = 2
            elif pct < 0.75:
                heatmap[d] = 3
            else:
                heatmap[d] = 4
    
    # --- Subject-Wise Stats ---
    subjects = db.query(models.Subject).filter(models.Subject.user_id == user_id).all()
    all_attendances = db.query(models.Attendance).filter(models.Attendance.user_id == user_id).all()
    
    subject_stats = {}
    for sub in subjects:
        sub_atts = [a for a in all_attendances if a.subject_id == sub.id]
        sub_map = {sub.id: sub}
        p = sum(_get_attendance_weight(a, sub_map) for a in sub_atts if a.status.lower() in ("present", "late_entry", "od", "event_leave", "medical_leave"))
        ab = sum(_get_attendance_weight(a, sub_map) for a in sub_atts if a.status.lower() == "absent")
        total = p + ab
        pct = round(p / total * 100, 1) if total > 0 else 0.0
        subject_stats[sub.name] = {
            "name": sub.name,
            "code": sub.code,
            "color": sub.color,
            "prof": sub.prof,
            "present": p,
            "absent": ab,
            "total": total,
            "percent": pct
        }
    
    return {
        "weekly_trend": weekly_trend,
        "heatmap": heatmap,
        "subjects": subject_stats
    }

@app.get("/analytics/prediction", tags=["Analytics"])
def get_prediction(
    missed: int = 0,
    attended: int = 0,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    """
    Prediction engine:
    - missed: how many classes the user plans to miss
    - attended: how many classes the user plans to attend
    Returns the predicted attendance percentage after those classes.
    """
    stats = _compute_global_stats(current_user.id, db)
    target = current_user.attendance_goal
    target_fraction = target / 100.0
    
    new_present = stats["present"] + attended
    new_total = stats["total"] + missed + attended
    
    predicted_pct = round(new_present / new_total * 100, 2) if new_total > 0 else 0.0
    
    # How many classes needed to reach target from current state
    if stats["total"] > 0:
        classes_for_target = max(0, math.ceil((target_fraction * stats["total"] - stats["present"]) / (1 - target_fraction)))
    else:
        classes_for_target = 0
    
    # Safe bunks from current state
    safe_bunks = max(0, math.floor((stats["present"] - target_fraction * stats["total"]) / target_fraction))
    
    return {
        "current_percent": stats["percentage"],
        "predicted_percent": predicted_pct,
        "will_reach_target": predicted_pct >= target,
        "classes_for_target": classes_for_target if stats["percentage"] < target else 0,
        "safe_bunks": safe_bunks if stats["percentage"] >= target else 0,
        "target": target,
        "scenarios": {
            "if_miss_1": round(stats["present"] / (stats["total"] + 1) * 100, 1) if stats["total"] > 0 else 0,
            "if_miss_3": round(stats["present"] / (stats["total"] + 3) * 100, 1) if stats["total"] > 0 else 0,
            "if_miss_5": round(stats["present"] / (stats["total"] + 5) * 100, 1) if stats["total"] > 0 else 0,
            "if_attend_5": round((stats["present"] + 5) / (stats["total"] + 5) * 100, 1) if stats["total"] > 0 else 100,
            "if_attend_10": round((stats["present"] + 10) / (stats["total"] + 10) * 100, 1) if stats["total"] > 0 else 100,
        }
    }

# --- Leave Plans Endpoints ---
@app.post("/leave_plans", response_model=schemas.LeavePlan, tags=["Leave Plans"])
def create_leave_plan(plan: schemas.LeavePlanCreate, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    db_plan = models.LeavePlan(**plan.model_dump(), user_id=current_user.id)
    db.add(db_plan)
    db.commit()
    db.refresh(db_plan)
    return db_plan

@app.get("/leave_plans", response_model=List[schemas.LeavePlan], tags=["Leave Plans"])
def get_leave_plans(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    return db.query(models.LeavePlan).filter(models.LeavePlan.user_id == current_user.id).order_by(models.LeavePlan.start_date.asc()).all()

@app.delete("/leave_plans/{plan_id}", tags=["Leave Plans"])
def delete_leave_plan(plan_id: int, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    db_plan = db.query(models.LeavePlan).filter(models.LeavePlan.id == plan_id, models.LeavePlan.user_id == current_user.id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="Leave plan not found")
    db.delete(db_plan)
    db.commit()
    return {"message": "Leave plan deleted successfully"}

# --- Profile Updates ---
@app.post("/user/device-token", tags=["App"])
def register_device_token(payload: schemas.FCMTokenRequest, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    current_user.fcm_token = payload.token
    db.commit()
    return {"message": "Device token registered"}

@app.put("/user/profile", tags=["App"])
def update_profile(profile: schemas.UserUpdate, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    current_user.name = profile.name
    current_user.attendance_goal = profile.attendance_goal
    current_user.semester = profile.semester
    current_user.college = profile.college
    current_user.branch = profile.branch
    current_user.roll_number = profile.roll_number
    current_user.section = profile.section
    current_user.year = profile.year
    current_user.register_number = profile.register_number
    current_user.university = profile.university
    current_user.profile_photo = profile.profile_photo
    db.commit()
    db.refresh(current_user)
    return {"message": "Profile updated successfully"}

# --- Change Password ---
@app.post("/auth/change-password", tags=["Authentication"])
def change_password(req: schemas.ChangePasswordRequest, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    if not auth.verify_password(req.current_password, current_user.password_hash):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    current_user.password_hash = auth.get_password_hash(req.new_password)
    db.commit()
    return {"message": "Password changed successfully"}

# --- Forgot Password (console-logged token for dev; swap for email in production) ---
@app.post("/auth/forgot-password", tags=["Authentication"])
def forgot_password(body: dict, db: Session = Depends(get_db)):
    from datetime import datetime
    email = body.get("email", "").strip().lower()
    user = db.query(models.User).filter(models.User.email == email).first()
    # Always return 200 to prevent email enumeration
    if user:
        raw_token = secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
        expires_at = datetime.utcnow() + timedelta(minutes=15)
        _reset_tokens[token_hash] = {"email": email, "expires_at": expires_at}
        # In production: send email. In dev: log to console.
        print(f"\n[DEV] Password reset token for {email}: {raw_token}\n")
    return {"message": "If that email is registered, a reset link has been sent (check server console in dev mode)"}

@app.post("/auth/reset-password", tags=["Authentication"])
def reset_password(body: dict, db: Session = Depends(get_db)):
    from datetime import datetime
    token = body.get("token", "").strip()
    new_password = body.get("new_password", "")
    if len(new_password) < 8 or not any(c.isdigit() for c in new_password):
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters and contain a number")
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    entry = _reset_tokens.get(token_hash)
    if not entry:
        raise HTTPException(status_code=400, detail="Invalid or expired reset token")
    if datetime.utcnow() > entry["expires_at"]:
        del _reset_tokens[token_hash]
        raise HTTPException(status_code=400, detail="Reset token has expired")
    user = db.query(models.User).filter(models.User.email == entry["email"]).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.password_hash = auth.get_password_hash(new_password)
    db.commit()
    del _reset_tokens[token_hash]
    return {"message": "Password reset successfully"}

# --- Timetable Sync ---
@app.post("/timetable/sync", tags=["App"])
def sync_timetable(req: schemas.TimetableSyncRequest, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    user_id = current_user.id
    
    # 1. Clear existing timetable entries
    db.query(models.Timetable).filter(models.Timetable.user_id == user_id).delete()
    db.commit()
    
    # 2. Add new timetable entries, resolving subjects by name
    from datetime import time as dt_time
    import random
    
    existing_subjects = db.query(models.Subject).filter(models.Subject.user_id == user_id).all()
    sub_map = {s.name.lower().strip(): s for s in existing_subjects}
    
    for entry in req.timetable:
        sub_key = entry.subject.lower().strip()
        # Ignore breaks when saving to database
        if entry.type == "Break" or sub_key in ["break", "lunch break", "recess", "lunch"]:
            continue
            
        if sub_key in sub_map:
            subject = sub_map[sub_key]
        else:
            colors = ["#cdbdff", "#40e56c", "#ffb3ae", "#7c4dff", "#02c953", "#ffdad7"]
            color = random.choice(colors)
            subject = models.Subject(
                user_id=user_id,
                name=entry.subject,
                code="CS-" + str(random.randint(100, 999)),
                prof=entry.prof,
                color=color,
                subject_type="Practical" if "lab" in entry.subject.lower() or "practical" in entry.subject.lower() or entry.type.lower() == "practical" else "Theory",
                credits=3
            )
            db.add(subject)
            db.commit()
            db.refresh(subject)
            sub_map[sub_key] = subject
            
        start_parts = [int(p) for p in entry.start.split(":")]
        end_parts = [int(p) for p in entry.end.split(":")]
        start_time = dt_time(start_parts[0], start_parts[1])
        end_time = dt_time(end_parts[0], end_parts[1])
        
        db_tt = models.Timetable(
            user_id=user_id,
            subject_id=subject.id,
            day=entry.day,
            start_time=start_time,
            end_time=end_time,
            room=entry.room,
            type=entry.type
        )
        db.add(db_tt)
        
    db.commit()
    
    # Regenerate class sessions for active semester with the new timetable
    _generate_sessions_for_active_semester(db, user_id)
    
    return {"message": "Timetable synced successfully"}

# --- Reports Endpoints ---
@app.get("/reports/summary", tags=["Reports"])
def get_report_summary(
    period: str = "monthly",
    start_date: str = None,
    end_date: str = None,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    """
    Returns structured data for report generation.
    period: daily, weekly, monthly, semester, custom
    """
    user_id = current_user.id
    today = date.today()
    
    # Determine date range
    if period == "daily":
        d_start = today
        d_end = today
    elif period == "weekly":
        d_start = today - timedelta(days=today.weekday())
        d_end = d_start + timedelta(days=6)
    elif period == "monthly":
        d_start = today.replace(day=1)
        next_month = today.replace(day=28) + timedelta(days=4)
        d_end = next_month - timedelta(days=next_month.day)
    elif period == "semester":
        d_start = today - timedelta(days=120)
        d_end = today
    elif period == "custom" and start_date and end_date:
        d_start = date.fromisoformat(start_date)
        d_end = date.fromisoformat(end_date)
    else:
        d_start = today - timedelta(days=30)
        d_end = today
    
    attendances = db.query(models.Attendance).filter(
        models.Attendance.user_id == user_id,
        models.Attendance.date >= d_start,
        models.Attendance.date <= d_end
    ).all()
    
    subjects = db.query(models.Subject).filter(models.Subject.user_id == user_id).all()
    sub_map = {s.id: s for s in subjects}
    
    # Overall stats
    present = sum(1 for a in attendances if a.status.lower() in ("present", "late_entry", "od", "event_leave", "medical_leave"))
    absent = sum(1 for a in attendances if a.status.lower() == "absent")
    cancelled = sum(1 for a in attendances if a.status.lower() == "cancelled")
    holidays = sum(1 for a in attendances if a.status.lower() == "holiday")
    total = present + absent
    percentage = round((present / total * 100), 2) if total > 0 else 0.0
    
    # Subject-wise breakdown
    subject_breakdown = []
    for sub in subjects:
        sub_atts = [a for a in attendances if a.subject_id == sub.id]
        s_present = sum(1 for a in sub_atts if a.status.lower() in ("present", "late_entry", "od", "event_leave", "medical_leave"))
        s_absent = sum(1 for a in sub_atts if a.status.lower() == "absent")
        s_total = s_present + s_absent
        s_pct = round((s_present / s_total * 100), 2) if s_total > 0 else 0.0
        subject_breakdown.append({
            "name": sub.name,
            "code": sub.code,
            "present": s_present,
            "absent": s_absent,
            "total": s_total,
            "percentage": s_pct,
            "color": sub.color
        })
    
    # Daily log for the range
    daily_log = {}
    for a in attendances:
        d_str = a.date.isoformat()
        if d_str not in daily_log:
            daily_log[d_str] = []
        daily_log[d_str].append({
            "subject": sub_map[a.subject_id].name if a.subject_id in sub_map else "Unknown",
            "status": a.status
        })
    
    return {
        "period": period,
        "start_date": d_start.isoformat(),
        "end_date": d_end.isoformat(),
        "student": {
            "name": current_user.name,
            "college": current_user.college,
            "branch": current_user.branch,
            "semester": current_user.semester,
            "target_goal": current_user.attendance_goal
        },
        "overall": {
            "present": present,
            "absent": absent,
            "cancelled": cancelled,
            "holidays": holidays,
            "total_conducted": total,
            "percentage": percentage
        },
        "subjects": subject_breakdown,
        "daily_log": daily_log
    }

# --- AttendWise App Specific Endpoints ---

from pydantic import BaseModel

class MarkAttendanceRequest(BaseModel):
    date: str
    subject_name: str
    start: str
    status: str

@app.post("/attendance/mark", tags=["App"])
def mark_attendance(req: MarkAttendanceRequest, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    user_id = current_user.id
    subject = db.query(models.Subject).filter(models.Subject.name == req.subject_name, models.Subject.user_id == user_id).first()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")
        
    date_obj = date.fromisoformat(req.date)
    
    # Parse start time from HH:MM string
    try:
        h, m = map(int, req.start.split(':'))
        start_time_obj = time(h, m)
    except Exception:
        start_time_obj = None
    
    # Find existing record for this subject on this day for the specific session
    att = db.query(models.Attendance).filter(
        models.Attendance.user_id == user_id, 
        models.Attendance.subject_id == subject.id,
        models.Attendance.date == date_obj,
        models.Attendance.start_time == start_time_obj
    ).first()
    
    # Sync with ClassSession if it exists
    session = db.query(models.ClassSession).filter(
        models.ClassSession.user_id == user_id,
        models.ClassSession.subject_id == subject.id,
        models.ClassSession.date == date_obj,
        models.ClassSession.start_time == start_time_obj
    ).first()
    
    if session:
        session.status = req.status
        
    if att:
        att.status = req.status
        if session:
            att.session_id = session.id
    else:
        att = models.Attendance(
            user_id=user_id,
            subject_id=subject.id,
            date=date_obj,
            start_time=start_time_obj,
            status=req.status,
            session_id=session.id if session else None
        )
        db.add(att)
    db.commit()
    return {"message": "Success"}

@app.get("/state", tags=["App"])
def get_state(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    user_id = current_user.id
    user = current_user
        
    subjects = db.query(models.Subject).filter(models.Subject.user_id == user_id).all()
    timetable = db.query(models.Timetable).filter(models.Timetable.user_id == user_id).all()
    attendances = db.query(models.Attendance).filter(models.Attendance.user_id == user_id).all()
    active_semester = db.query(models.Semester).filter(
        models.Semester.user_id == user_id,
        models.Semester.is_active == True
    ).first()
    
    # Compute streak
    streak = _compute_streak(user_id, db)
    
    # Compute global stats for safe_bunks
    stats = _compute_global_stats(user_id, db)
    bunk_info = _compute_safe_bunks(stats["present"], stats["total"], user.attendance_goal)
    
    sub_map = {s.id: s for s in subjects}
    tt_formatted = []
    for t in timetable:
        tt_formatted.append({
            "day": t.day,
            "subject": sub_map[t.subject_id].name if t.subject_id in sub_map else "Unknown",
            "start": t.start_time.strftime("%H:%M"),
            "end": t.end_time.strftime("%H:%M"),
            "room": t.room,
            "prof": sub_map[t.subject_id].prof if t.subject_id in sub_map else None,
            "type": t.type,
            "color": sub_map[t.subject_id].color if t.subject_id in sub_map else "#7c4dff"
        })
        
    logs = {}
    days_map = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for a in attendances:
        d_str = a.date.isoformat()
        if d_str not in logs:
            logs[d_str] = []
        
        day_str = days_map[a.date.weekday()]
        
        # Match using subject AND start_time (if available) to find the correct timetable entry for end_time
        tt_entry = next((t for t in timetable if t.day == day_str and t.subject_id == a.subject_id and t.start_time == a.start_time), None)
        
        start_str = a.start_time.strftime("%H:%M") if a.start_time else "00:00"
        end_str = tt_entry.end_time.strftime("%H:%M") if tt_entry else "00:00"
        
        logs[d_str].append({
            "subject": sub_map[a.subject_id].name if a.subject_id in sub_map else "Unknown",
            "start": start_str,
            "end": end_str,
            "status": a.status,
            "color": sub_map[a.subject_id].color if a.subject_id in sub_map else "#7c4dff"
        })

    holidays = db.query(models.Holiday).filter(models.Holiday.user_id == user_id).all()
    holidays_data = [{"date": h.date.isoformat(), "name": h.name, "type": h.type} for h in holidays]
    return {
        "holidays": holidays_data,
        "profile": {
            "name": user.name,
            "email": user.email,
            "targetGoal": user.attendance_goal,
            "term": user.semester,
            "streak": streak,
            "college": user.college,
            "branch": user.branch,
            "roll_number": user.roll_number,
            "section": user.section,
            "year": user.year,
            "register_number": user.register_number,
            "university": user.university,
            "profile_photo": user.profile_photo,
        },
        "active_semester": {
            "id": active_semester.id,
            "name": active_semester.name,
            "start_date": active_semester.start_date.isoformat() if active_semester.start_date else None,
            "end_date": active_semester.end_date.isoformat() if active_semester.end_date else None,
            "academic_year": active_semester.academic_year,
            "academic_calendar": active_semester.academic_calendar
        } if active_semester else None,
        "globalStats": {
            "percentage": stats["percentage"],
            "present": stats["present"],
            "absent": stats["absent"],
            "total": stats["total"]
        },
        "bunkAnalysis": bunk_info,
        "subjects": [{"id": s.id, "name": s.name, "code": s.code, "prof": s.prof, "color": s.color, "credits": s.credits, "minimum_required_attendance": s.minimum_required_attendance} for s in subjects],
        "timetable": tt_formatted,
        "attendanceLogs": logs
    }

# --- Missing / Calendar-Driven Endpoints ---

@app.get("/user/profile", response_model=schemas.User, tags=["User"])
def get_user_profile(current_user: models.User = Depends(auth.get_current_user)):
    return current_user

@app.post("/auth/logout", tags=["Authentication"])
def logout():
    return {"message": "Successfully logged out"}

@app.get("/analytics/streak", tags=["Analytics"])
def get_streak(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    streak = _compute_streak(current_user.id, db)
    return {"streak": streak}

def _normalize_date_str(s: str) -> str:
    """Normalize DD-MM-YYYY or D-M-YYYY to YYYY-MM-DD for fromisoformat."""
    if not s:
        return s
    s = s.strip()
    import re as _re
    m = _re.match(r'^(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})$', s)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    return s

def _expand_date_range(start_str: str, end_str: str) -> List[date]:
    try:
        start_d = date.fromisoformat(_normalize_date_str(start_str))
        end_d = date.fromisoformat(_normalize_date_str(end_str))
        curr = start_d
        res = []
        while curr <= end_d:
            res.append(curr)
            curr += timedelta(days=1)
        return res
    except Exception:
        return []

@app.post("/semesters", response_model=schemas.SemesterOut, tags=["Semesters"])
def create_semester(
    sem: schemas.SemesterCreate,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    # Capture plain integers BEFORE any commit (ORM objects get expired after commit)
    user_id = int(current_user.id)

    db.query(models.Semester).filter(models.Semester.user_id == user_id).update({"is_active": False})

    db_sem = models.Semester(
        user_id=user_id,
        name=sem.name,
        academic_year=sem.academic_year,
        start_date=sem.start_date,
        end_date=sem.end_date,
        academic_calendar=sem.academic_calendar,
        is_active=True
    )
    db.add(db_sem)
    db.commit()
    db.refresh(db_sem)

    # Capture sem_id as plain integer after refresh
    sem_id = int(db_sem.id)

    # Populate Holidays using raw SQL INSERT OR IGNORE to safely skip duplicates
    if sem.academic_calendar:
        try:
            cal = json.loads(sem.academic_calendar)

            # Collect all (date, name, type) tuples to insert
            rows = []

            def _collect(hdate, hname, htype):
                if hdate:
                    rows.append((str(hdate), hname, htype))

            for h in cal.get("holidays", []):
                try:
                    _collect(date.fromisoformat(h["date"]), h.get("name", "Holiday"), "Holiday")
                except Exception:
                    pass

            for m in cal.get("midExams", []):
                for d in _expand_date_range(m.get("start", ""), m.get("end", "")):
                    _collect(d, m.get("title", "Mid Exams"), "Mid Exam")

            for l in cal.get("labExams", []):
                for d in _expand_date_range(l.get("start", ""), l.get("end", "")):
                    _collect(d, l.get("title", "Lab Exams"), "Lab Exam")

            for b in cal.get("semesterBreak", []):
                for d in _expand_date_range(b.get("start", ""), b.get("end", "")):
                    _collect(d, b.get("title", "Semester Break"), "Semester Break")

            for e in cal.get("examDates", []):
                for d in _expand_date_range(e.get("start", ""), e.get("end", "")):
                    _collect(d, e.get("title", "Semester Exams"), "Semester Exam")

            for s in cal.get("studyHolidays", []):
                for d in _expand_date_range(s.get("start", ""), s.get("end", "")):
                    _collect(d, s.get("title", "Preparation Leave"), "Study Holiday")

            for ev in cal.get("events", []):
                try:
                    _collect(date.fromisoformat(ev["date"]), ev.get("title", "College Event"), "Event")
                except Exception:
                    pass

            # Bulk insert with raw SQL INSERT OR IGNORE to skip any duplicates silently
            if rows:
                db.execute(
                    text(
                        "INSERT OR IGNORE INTO holidays (user_id, semester_id, date, name, type) "
                        "VALUES (:uid, :sid, :dt, :nm, :tp)"
                    ),
                    [{"uid": user_id, "sid": sem_id, "dt": r[0], "nm": r[1], "tp": r[2]} for r in rows]
                )
                db.commit()

        except Exception as err:
            db.rollback()
            print("Warning: Could not auto-populate holidays:", err)

    # Generate scheduled class sessions
    try:
        _generate_sessions_for_active_semester(db, user_id)
    except Exception as sess_err:
        db.rollback()
        print("Warning: Could not auto-generate sessions:", sess_err)

    # Re-fetch db_sem so it's in a clean state to return
    db_sem = db.query(models.Semester).filter(models.Semester.id == sem_id).first()
    return db_sem


@app.get("/semesters", response_model=List[schemas.SemesterOut], tags=["Semesters"])
def get_semesters(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    return db.query(models.Semester).filter(models.Semester.user_id == current_user.id).all()

@app.get("/semesters/{semester_id}", response_model=schemas.SemesterOut, tags=["Semesters"])
def get_semester(
    semester_id: int,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    sem = db.query(models.Semester).filter(
        models.Semester.id == semester_id,
        models.Semester.user_id == current_user.id
    ).first()
    if not sem:
        raise HTTPException(status_code=404, detail="Semester not found")
    return sem

@app.put("/semesters/{semester_id}/activate", response_model=schemas.SemesterOut, tags=["Semesters"])
def activate_semester(
    semester_id: int,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    sem = db.query(models.Semester).filter(
        models.Semester.id == semester_id,
        models.Semester.user_id == current_user.id
    ).first()
    if not sem:
        raise HTTPException(status_code=404, detail="Semester not found")
    
    db.query(models.Semester).filter(models.Semester.user_id == current_user.id).update({"is_active": False})
    sem.is_active = True
    db.commit()
    db.refresh(sem)
    
    # Generate scheduled class sessions for the newly active semester
    _generate_sessions_for_active_semester(db, current_user.id)
    
    return sem

@app.delete("/semesters/{semester_id}", tags=["Semesters"])
def delete_semester(
    semester_id: int,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    sem = db.query(models.Semester).filter(
        models.Semester.id == semester_id,
        models.Semester.user_id == current_user.id
    ).first()
    if not sem:
        raise HTTPException(status_code=404, detail="Semester not found")
    db.delete(sem)
    db.commit()
    return {"message": "Semester and all associated data deleted successfully"}

@app.post("/semesters/{semester_id}/holidays", response_model=schemas.HolidayOut, tags=["Holidays"])
def create_holiday(
    semester_id: int,
    holiday: schemas.HolidayCreate,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    sem = db.query(models.Semester).filter(
        models.Semester.id == semester_id,
        models.Semester.user_id == current_user.id
    ).first()
    if not sem:
        raise HTTPException(status_code=404, detail="Semester not found")
        
    db_holiday = db.query(models.Holiday).filter(
        models.Holiday.user_id == current_user.id,
        models.Holiday.semester_id == semester_id,
        models.Holiday.date == holiday.date
    ).first()
    if db_holiday:
        raise HTTPException(status_code=400, detail="Holiday already exists on this date")
        
    db_holiday = models.Holiday(
        user_id=current_user.id,
        semester_id=semester_id,
        date=holiday.date,
        name=holiday.name,
        type=holiday.type
    )
    db.add(db_holiday)
    
    # Mark any class sessions on this date as holiday status
    db.query(models.ClassSession).filter(
        models.ClassSession.user_id == current_user.id,
        models.ClassSession.semester_id == semester_id,
        models.ClassSession.date == holiday.date,
        models.ClassSession.status == "upcoming"
    ).update({"status": "holiday"})
    
    db.commit()
    db.refresh(db_holiday)
    return db_holiday

@app.get("/semesters/{semester_id}/holidays", response_model=List[schemas.HolidayOut], tags=["Holidays"])
def get_holidays(
    semester_id: int,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    sem = db.query(models.Semester).filter(
        models.Semester.id == semester_id,
        models.Semester.user_id == current_user.id
    ).first()
    if not sem:
        raise HTTPException(status_code=404, detail="Semester not found")
    return db.query(models.Holiday).filter(
        models.Holiday.user_id == current_user.id,
        models.Holiday.semester_id == semester_id
    ).all()

@app.delete("/semesters/{semester_id}/holidays/{holiday_id}", tags=["Holidays"])
def delete_holiday(
    semester_id: int,
    holiday_id: int,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    h = db.query(models.Holiday).filter(
        models.Holiday.id == holiday_id,
        models.Holiday.semester_id == semester_id,
        models.Holiday.user_id == current_user.id
    ).first()
    if not h:
        raise HTTPException(status_code=404, detail="Holiday not found")
        
    holiday_date = h.date
    db.delete(h)
    
    # Revert any holiday class sessions on this date back to upcoming
    db.query(models.ClassSession).filter(
        models.ClassSession.user_id == current_user.id,
        models.ClassSession.semester_id == semester_id,
        models.ClassSession.date == holiday_date,
        models.ClassSession.status == "holiday"
    ).update({"status": "upcoming"})
    
    db.commit()
    return {"message": "Holiday deleted successfully"}

def _generate_sessions_for_active_semester(db: Session, user_id: int):
    semester = db.query(models.Semester).filter(
        models.Semester.user_id == user_id,
        models.Semester.is_active == True
    ).first()
    if not semester:
        return
        
    start_gen = semester.start_date
    end_gen = semester.end_date
    
    # Delete existing upcoming class sessions for this active semester
    db.query(models.ClassSession).filter(
        models.ClassSession.user_id == user_id,
        models.ClassSession.semester_id == semester.id,
        models.ClassSession.date >= start_gen,
        models.ClassSession.status == "upcoming"
    ).delete()
    db.commit()
    
    holidays = db.query(models.Holiday).filter(
        models.Holiday.user_id == user_id,
        models.Holiday.semester_id == semester.id
    ).all()
    holiday_dates = {h.date for h in holidays}
    
    # Parse working Saturdays to override holiday check
    working_saturdays = set()
    if semester.academic_calendar:
        try:
            cal = json.loads(semester.academic_calendar)
            working_saturdays = {date.fromisoformat(d) for d in cal.get("workingSaturdays", [])}
        except Exception:
            pass
    
    days_map = {0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"}
    
    # Load user's active timetable entries where version_id is None
    entries = db.query(models.Timetable).filter(
        models.Timetable.user_id == user_id,
        models.Timetable.version_id == None
    ).all()
    
    if not entries:
        # Fallback to load any timetable entries for the user if version_id is not used
        entries = db.query(models.Timetable).filter(models.Timetable.user_id == user_id).all()
        
    from collections import defaultdict
    day_entries = defaultdict(list)
    for entry in entries:
        day_entries[entry.day].append(entry)
        
    curr = start_gen
    to_add = []
    while curr <= end_gen:
        day_str = days_map[curr.weekday()]
        status = "holiday" if (curr in holiday_dates and curr not in working_saturdays) else "upcoming"
        
        if day_str == "Sun":
            curr += timedelta(days=1)
            continue
            
        for entry in day_entries[day_str]:
            # Check if there's already a session on this date and time
            session_exists = db.query(models.ClassSession).filter(
                models.ClassSession.user_id == user_id,
                models.ClassSession.subject_id == entry.subject_id,
                models.ClassSession.date == curr,
                models.ClassSession.start_time == entry.start_time
            ).first()
            if not session_exists:
                to_add.append(models.ClassSession(
                    user_id=user_id,
                    semester_id=semester.id,
                    subject_id=entry.subject_id,
                    date=curr,
                    start_time=entry.start_time,
                    end_time=entry.end_time,
                    room=entry.room,
                    session_type=entry.type or "Lecture",
                    status=status,
                    is_extra=False
                ))
        curr += timedelta(days=1)
        
    if to_add:
        db.add_all(to_add)
        db.commit()

def _generate_sessions_for_version(db: Session, user_id: int, version: models.TimetableVersion):
    semester = db.query(models.Semester).filter(models.Semester.id == version.semester_id, models.Semester.user_id == user_id).first()
    if not semester:
        return
        
    start_gen = max(version.effective_from, semester.start_date)
    end_gen = semester.end_date
    
    # Delete existing upcoming class sessions on or after start_gen
    db.query(models.ClassSession).filter(
        models.ClassSession.user_id == user_id,
        models.ClassSession.semester_id == semester.id,
        models.ClassSession.date >= start_gen,
        models.ClassSession.status == "upcoming"
    ).delete()
    db.commit()
    
    holidays = db.query(models.Holiday).filter(
        models.Holiday.user_id == user_id,
        models.Holiday.semester_id == semester.id
    ).all()
    holiday_dates = {h.date for h in holidays}
    
    days_map = {0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"}
    
    entries = db.query(models.Timetable).filter(models.Timetable.version_id == version.id).all()
    
    from collections import defaultdict
    day_entries = defaultdict(list)
    for entry in entries:
        day_entries[entry.day].append(entry)
        
    curr = start_gen
    to_add = []
    while curr <= end_gen:
        day_str = days_map[curr.weekday()]
        status = "holiday" if curr in holiday_dates else "upcoming"
        
        for entry in day_entries[day_str]:
            # check if there's already a session on this date and time
            session_exists = db.query(models.ClassSession).filter(
                models.ClassSession.user_id == user_id,
                models.ClassSession.subject_id == entry.subject_id,
                models.ClassSession.date == curr,
                models.ClassSession.start_time == entry.start_time
            ).first()
            if not session_exists:
                to_add.append(models.ClassSession(
                    user_id=user_id,
                    semester_id=semester.id,
                    subject_id=entry.subject_id,
                    date=curr,
                    start_time=entry.start_time,
                    end_time=entry.end_time,
                    room=entry.room,
                    session_type=entry.type or "Lecture",
                    status=status,
                    is_extra=False
                ))
        curr += timedelta(days=1)
        
    if to_add:
        db.add_all(to_add)
        db.commit()

@app.post("/timetable/versions", response_model=schemas.TimetableVersionOut, tags=["Timetable Versions"])
def create_timetable_version(
    version: schemas.TimetableVersionCreate,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    sem = db.query(models.Semester).filter(
        models.Semester.id == version.semester_id,
        models.Semester.user_id == current_user.id
    ).first()
    if not sem:
        raise HTTPException(status_code=404, detail="Semester not found")
        
    db_version = models.TimetableVersion(
        user_id=current_user.id,
        semester_id=version.semester_id,
        label=version.label,
        effective_from=version.effective_from
    )
    db.add(db_version)
    db.commit()
    db.refresh(db_version)
    
    import random
    from datetime import time as dt_time
    
    existing_subjects = db.query(models.Subject).filter(models.Subject.user_id == current_user.id).all()
    sub_map = {s.name.lower().strip(): s for s in existing_subjects}
    
    for entry in version.timetable:
        sub_key = entry.subject.lower().strip()
        if sub_key in sub_map:
            subject = sub_map[sub_key]
        else:
            colors = ["#cdbdff", "#40e56c", "#ffb3ae", "#7c4dff", "#02c953", "#ffdad7"]
            color = random.choice(colors)
            subject = models.Subject(
                user_id=current_user.id,
                name=entry.subject,
                code="CS-" + str(random.randint(100, 999)),
                prof=entry.prof,
                color=color,
                subject_type="Practical" if "lab" in entry.subject.lower() or "practical" in entry.subject.lower() or entry.type.lower() == "practical" else "Theory",
                credits=3
            )
            db.add(subject)
            db.commit()
            db.refresh(subject)
            sub_map[sub_key] = subject
            
        start_parts = [int(p) for p in entry.start.split(":")]
        end_parts = [int(p) for p in entry.end.split(":")]
        start_time = dt_time(start_parts[0], start_parts[1])
        end_time = dt_time(end_parts[0], end_parts[1])
        
        db_tt = models.Timetable(
            user_id=current_user.id,
            subject_id=subject.id,
            day=entry.day,
            start_time=start_time,
            end_time=end_time,
            room=entry.room,
            type=entry.type,
            version_id=db_version.id
        )
        db.add(db_tt)
        
    db.commit()
    
    _generate_sessions_for_version(db, current_user.id, db_version)
    
    return db_version

@app.get("/timetable/versions", response_model=List[schemas.TimetableVersionOut], tags=["Timetable Versions"])
def get_timetable_versions(
    semester_id: int,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    return db.query(models.TimetableVersion).filter(
        models.TimetableVersion.user_id == current_user.id,
        models.TimetableVersion.semester_id == semester_id
    ).all()

@app.get("/sessions", response_model=List[schemas.ClassSessionOut], tags=["Class Sessions"])
def get_class_sessions(
    start_date: date,
    end_date: date,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    sessions = db.query(models.ClassSession).filter(
        models.ClassSession.user_id == current_user.id,
        models.ClassSession.date >= start_date,
        models.ClassSession.date <= end_date
    ).order_by(models.ClassSession.date.asc(), models.ClassSession.start_time.asc()).all()
    
    out = []
    for s in sessions:
        out.append(schemas.ClassSessionOut(
            id=s.id,
            date=s.date,
            start_time=s.start_time,
            end_time=s.end_time,
            room=s.room,
            session_type=s.session_type,
            status=s.status,
            is_extra=s.is_extra,
            subject_id=s.subject_id,
            subject_name=s.subject.name,
            subject_color=s.subject.color,
            subject_prof=s.subject.prof
        ))
    return out

@app.post("/sessions/{session_id}/mark", response_model=schemas.ClassSessionOut, tags=["Class Sessions"])
def mark_class_session(
    session_id: int,
    req: schemas.MarkSessionRequest,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    s = db.query(models.ClassSession).filter(
        models.ClassSession.id == session_id,
        models.ClassSession.user_id == current_user.id
    ).first()
    if not s:
        raise HTTPException(status_code=404, detail="Class session not found")
        
    s.status = req.status
    
    att = db.query(models.Attendance).filter(
        models.Attendance.session_id == s.id,
        models.Attendance.user_id == current_user.id
    ).first()
    
    if not att:
        att = db.query(models.Attendance).filter(
            models.Attendance.user_id == current_user.id,
            models.Attendance.subject_id == s.subject_id,
            models.Attendance.date == s.date
        ).first()
        
    if att:
        att.status = req.status
        att.remarks = req.remarks
        att.session_id = s.id
    else:
        att = models.Attendance(
            user_id=current_user.id,
            subject_id=s.subject_id,
            date=s.date,
            status=req.status,
            remarks=req.remarks,
            session_id=s.id
        )
        db.add(att)
        
    db.commit()
    db.refresh(s)
    
    return schemas.ClassSessionOut(
        id=s.id,
        date=s.date,
        start_time=s.start_time,
        end_time=s.end_time,
        room=s.room,
        session_type=s.session_type,
        status=s.status,
        is_extra=s.is_extra,
        subject_id=s.subject_id,
        subject_name=s.subject.name,
        subject_color=s.subject.color,
        subject_prof=s.subject.prof
    )

@app.post("/sessions/extra", response_model=schemas.ClassSessionOut, tags=["Class Sessions"])
def create_extra_class_session(
    extra: schemas.ExtraClassCreate,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    sem = db.query(models.Semester).filter(
        models.Semester.id == extra.semester_id,
        models.Semester.user_id == current_user.id
    ).first()
    if not sem:
        raise HTTPException(status_code=404, detail="Semester not found")
        
    sub = db.query(models.Subject).filter(
        models.Subject.id == extra.subject_id,
        models.Subject.user_id == current_user.id
    ).first()
    if not sub:
        raise HTTPException(status_code=404, detail="Subject not found")
        
    session_exists = db.query(models.ClassSession).filter(
        models.ClassSession.user_id == current_user.id,
        models.ClassSession.subject_id == extra.subject_id,
        models.ClassSession.date == extra.date,
        models.ClassSession.start_time == extra.start_time
    ).first()
    if session_exists:
        raise HTTPException(status_code=400, detail="Class session already exists at this date and time")
        
    s = models.ClassSession(
        user_id=current_user.id,
        semester_id=extra.semester_id,
        subject_id=extra.subject_id,
        date=extra.date,
        start_time=extra.start_time,
        end_time=extra.end_time,
        room=extra.room,
        session_type=extra.session_type or "Extra",
        status="upcoming",
        is_extra=True
    )
    db.add(s)
    db.commit()
    db.refresh(s)
    
    return schemas.ClassSessionOut(
        id=s.id,
        date=s.date,
        start_time=s.start_time,
        end_time=s.end_time,
        room=s.room,
        session_type=s.session_type,
        status=s.status,
        is_extra=s.is_extra,
        subject_id=s.subject_id,
        subject_name=s.subject.name,
        subject_color=s.subject.color,
        subject_prof=s.subject.prof
    )

# ============================================================================
# AI OCR TIMETABLE PARSER (Gemini 2.5 Flash)
# ============================================================================

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
GEMINI_MODEL   = "gemini-2.5-flash"
GEMINI_URL     = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"

TIMETABLE_OCR_PROMPT = """You are an expert Indian engineering college timetable parser.
Analyze the provided timetable image and extract ALL class/lab entries for every day.

IMPORTANT RULES FOR INDIAN TIMETABLES:
1. Subject codes (abbreviations) like CO, CN, EDA, AI, OE must be resolved to full names using the subject legend/list table if present in the image. Always prefer the full name.
2. Lab sessions spanning multiple periods (e.g. "Tinkering Lab" across periods 2-3-4) should be ONE entry with start time = first period start, end time = last period end.
3. "Mentoring/Seminar", "Mentor", "Seminar" entries are valid classes — include them as type "Tutorial".
4. "BREAK" or "Recess" columns are lunch/tea breaks — include as type "Break" with subject "Break".
5. Period timings: If the image shows period numbers with times, map period → time carefully.

Return ONLY a valid JSON object (no markdown, no explanation) with this exact structure:
{
  "timetable": [
    {
      "day": "Mon",
      "subject": "Full Subject Name",
      "start": "09:00",
      "end": "09:50",
      "room": "",
      "prof": "",
      "type": "Lecture"
    }
  ],
  "total_classes": 12,
  "message": "Successfully extracted N classes"
}

Rules:
- day must be one of: Mon, Tue, Wed, Thu, Fri, Sat
- start and end must be in HH:MM 24-hour format (e.g. 09:00, 14:30)
- type must be one of: Lecture, Practical, Tutorial, Break
- If room or prof is not visible, use empty string ""
- If a cell is empty or "Free", skip it
- Resolve ALL subject abbreviations to full names using the legend table in the image
- For labs spanning multiple consecutive periods, create ONE entry covering the full time range (type: "Practical")
- Include every day including Saturday if present
"""

@app.post("/timetable/ocr", tags=["Timetable"])
async def ocr_timetable(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.get_current_user)
):
    """
    Upload a timetable image (PNG/JPEG/WebP) or PDF.
    Gemini AI will parse it and return a structured list of class entries.
    """
    api_key = os.getenv("GEMINI_API_KEY", GEMINI_API_KEY)
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail="Gemini API key not configured. Set GEMINI_API_KEY in your .env file."
        )

    # File size validation (max 10MB)
    file_bytes = await file.read()
    MAX_FILE_SIZE = 10 * 1024 * 1024
    if len(file_bytes) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File too large. Maximum size is 10MB.")

    # File type validation
    ALLOWED_MIME_TYPES = {"image/png", "image/jpeg", "image/webp", "image/gif", "application/pdf"}
    mime_type = file.content_type or "image/png"
    if file.filename and file.filename.lower().endswith(".pdf"):
        mime_type = "application/pdf"
    if mime_type not in ALLOWED_MIME_TYPES:
        raise HTTPException(status_code=415, detail=f"Unsupported file type: {mime_type}. Allowed: PNG, JPEG, WebP, PDF.")

    encoded = base64.b64encode(file_bytes).decode("utf-8")

    payload = {
        "contents": [
            {
                "parts": [
                    {"text": TIMETABLE_OCR_PROMPT},
                    {
                        "inlineData": {
                            "mimeType": mime_type,
                            "data": encoded
                        }
                    }
                ]
            }
        ],
        "generationConfig": {
            "response_mime_type": "application/json",
            "temperature": 0.1
        }
    }

    headers = {"Content-Type": "application/json"}
    
    models_to_try = ["gemini-3.5-flash", "gemini-3.1-flash-lite", "gemini-flash-latest", "gemini-2.5-flash-lite"]
    gemini_response = None
    last_error = None
    successful_model = None

    for model in models_to_try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
        try:
            import urllib.request as ureq
            req = ureq.Request(
                url,
                data=json.dumps(payload).encode("utf-8"),
                headers=headers,
                method="POST"
            )
            with ureq.urlopen(req, timeout=30) as resp:
                gemini_response = json.loads(resp.read().decode("utf-8"))
                successful_model = model
                break
        except Exception as e:
            error_msg = str(e)
            if hasattr(e, 'read'):
                try:
                    error_msg = e.read().decode("utf-8")
                except Exception:
                    pass
            print(f"Gemini API call failed for model {model}: {error_msg}")
            last_error = error_msg

    if not gemini_response:
        raise HTTPException(status_code=500, detail=f"Gemini OCR failed: {last_error}")

    try:
        candidates = gemini_response.get("candidates", [])
        if not candidates:
            raise HTTPException(status_code=500, detail="Gemini returned no candidates")

        raw_text = candidates[0]["content"]["parts"][0]["text"]
        parsed = json.loads(raw_text)
        timetable = parsed.get("timetable", [])

        valid_days = {"Mon", "Tue", "Wed", "Thu", "Fri", "Sat"}
        valid_types = {"Lecture", "Practical", "Hybrid", "Tutorial", "Break"}
        cleaned = []
        for entry in timetable:
            day = entry.get("day", "").strip()
            subject = entry.get("subject", "").strip()
            start = entry.get("start", "").strip()
            end = entry.get("end", "").strip()
            if day in valid_days and subject and start and end:
                cleaned.append({
                    "day": day,
                    "subject": subject,
                    "start": start,
                    "end": end,
                    "room": entry.get("room", "") or "",
                    "prof": entry.get("prof", "") or "",
                    "type": entry.get("type", "Lecture") if entry.get("type") in valid_types else "Lecture"
                })

        return {
            "timetable": cleaned,
            "total_classes": len(cleaned),
            "message": f"AI OCR extracted {len(cleaned)} classes from your timetable",
            "model": successful_model
        }

    except HTTPException:
        raise
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse Gemini JSON response: {str(e)}")
    except Exception as e:
        error_msg = str(e)
        if hasattr(e, 'read'):
            try:
                error_msg = e.read().decode("utf-8")
            except Exception:
                pass
        print("GEMINI ERROR:", error_msg)
        raise HTTPException(status_code=500, detail=f"Gemini OCR failed: {error_msg}")

ATTENDANCE_OCR_PROMPT = """You are an expert OCR parser for student portals.
Analyze the provided screenshot of the student attendance portal and extract the conducted and attended classes for all subjects.
Return ONLY a valid JSON list of objects (no markdown, no explanations) with this exact structure:
[
  {
    "subject_name": "Subject Name",
    "conducted": 42,
    "attended": 35
  }
]
- Include all subjects present in the image.
- `conducted` and `attended` MUST be integers.
- If you see a percentage instead of raw numbers, and there is no raw number, try to infer or just output what is available, but raw numbers are preferred.
"""

@app.post("/attendance/ocr", tags=["Attendance"])
async def ocr_attendance(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.get_current_user)
):
    """
    Upload a student portal screenshot.
    Gemini AI will parse it and return a list of subjects with conducted/attended counts.
    """
    api_key = os.getenv("GEMINI_API_KEY", GEMINI_API_KEY)
    if not api_key:
        raise HTTPException(status_code=503, detail="Gemini API key not configured.")

    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large. Maximum size is 10MB.")

    ALLOWED_MIME_TYPES = {"image/png", "image/jpeg", "image/webp", "image/gif", "application/pdf"}
    mime_type = file.content_type or "image/png"
    if file.filename and file.filename.lower().endswith(".pdf"):
        mime_type = "application/pdf"
    if mime_type not in ALLOWED_MIME_TYPES:
        mime_type = "image/png"
        
    import base64
    encoded = base64.b64encode(file_bytes).decode("utf-8")

    payload = {
        "contents": [
            {
                "parts": [
                    {"text": ATTENDANCE_OCR_PROMPT},
                    {
                        "inlineData": {
                            "mimeType": mime_type,
                            "data": encoded
                        }
                    }
                ]
            }
        ],
        "generationConfig": {
            "response_mime_type": "application/json",
            "temperature": 0.1
        }
    }

    headers = {"Content-Type": "application/json"}
    models_to_try = ["gemini-3.5-flash", "gemini-3.1-flash-lite", "gemini-flash-latest", "gemini-2.5-flash-lite"]
    gemini_response = None
    last_error = None

    for model in models_to_try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
        try:
            import urllib.request as ureq
            req = ureq.Request(url, data=json.dumps(payload).encode("utf-8"), headers=headers, method="POST")
            with ureq.urlopen(req, timeout=30) as resp:
                gemini_response = json.loads(resp.read().decode("utf-8"))
                break
        except Exception as e:
            last_error = str(e)
            if hasattr(e, 'read'):
                try: last_error = e.read().decode("utf-8")
                except Exception: pass

    if not gemini_response:
        raise HTTPException(status_code=500, detail=f"Gemini OCR failed: {last_error}")

    try:
        candidates = gemini_response.get("candidates", [])
        if not candidates:
            raise HTTPException(status_code=500, detail="Gemini returned no candidates")

        raw_text = candidates[0]["content"]["parts"][0]["text"]
        parsed = json.loads(raw_text)
        if not isinstance(parsed, list):
            raise ValueError("Expected a JSON list")
            
        return parsed
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse Gemini response: {str(e)}")


CALENDAR_OCR_PROMPT = """You are an expert Indian academic calendar parser for engineering colleges.
Analyze the provided academic calendar image or PDF and extract ALL key dates and events.

IMPORTANT: Always convert dates to YYYY-MM-DD format regardless of how they appear in the image (e.g. 29-06-2026 → 2026-06-29).

Return ONLY a valid JSON object (no markdown, no explanations) with this exact structure:
{
  "semesterStart": "YYYY-MM-DD",
  "semesterEnd": "YYYY-MM-DD",
  "holidays": [
    {"date": "YYYY-MM-DD", "name": "Name of holiday"}
  ],
  "midExams": [
    {"title": "I Mid Examinations", "start": "YYYY-MM-DD", "end": "YYYY-MM-DD"}
  ],
  "labExams": [
    {"title": "Lab Exams", "start": "YYYY-MM-DD", "end": "YYYY-MM-DD"}
  ],
  "semesterBreak": [
    {"title": "Dasara Holidays", "start": "YYYY-MM-DD", "end": "YYYY-MM-DD"}
  ],
  "examDates": [
    {"title": "End Examinations", "start": "YYYY-MM-DD", "end": "YYYY-MM-DD"}
  ],
  "studyHolidays": [
    {"title": "Practical Examinations & Preparation", "start": "YYYY-MM-DD", "end": "YYYY-MM-DD"}
  ],
  "workingSaturdays": ["YYYY-MM-DD"],
  "events": [
    {"title": "Sports Day", "date": "YYYY-MM-DD"}
  ]
}

Categorization Rules for Indian Engineering College Calendars:
- semesterStart: The "Commencement of Class Work" date
- semesterEnd: The last date of "End Examinations" / "Semester Examinations"
- midExams: Any row containing "Mid Examination", "Unit Test", "Internal Exam" — extract start and end dates
- labExams: "Lab Examination" or "Practical Examination" periods that are clearly lab/practical exams
- semesterBreak: Festival holidays like "Dasara", "Dussehra", "Puja Holidays", "Diwali Holidays", "Christmas Holidays", "Winter Break", "Summer Break"
- examDates: "End Examination", "Semester Examination", "Final Examination" periods
- studyHolidays: "Practical Examinations & Preparation", "Preparation Leave", "Study Holidays", "Pre-Exam Holiday" — these are days before end exams
- holidays: Individual public holidays, national holidays (Independence Day, Republic Day, etc.)
- workingSaturdays: Explicitly mentioned as "Special Working Saturday" or similar
- events: College fests, sports days, annual day events

All dates MUST be in YYYY-MM-DD format. Convert from DD-MM-YYYY if needed.
"""

@app.post("/semester/parse-calendar", tags=["Semesters"])
async def parse_calendar(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.get_current_user)
):
    """
    Upload an academic calendar image (PNG/JPEG/WebP) or PDF.
    Gemini AI will parse it and return a structured JSON configuration.
    """
    api_key = os.getenv("GEMINI_API_KEY", GEMINI_API_KEY)
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail="Gemini API key not configured. Set GEMINI_API_KEY in your .env file."
        )

    file_bytes = await file.read()
    encoded = base64.b64encode(file_bytes).decode("utf-8")

    mime_type = file.content_type or "image/png"
    if file.filename and file.filename.lower().endswith(".pdf"):
        mime_type = "application/pdf"

    payload = {
        "contents": [
            {
                "parts": [
                    {"text": CALENDAR_OCR_PROMPT},
                    {
                        "inlineData": {
                            "mimeType": mime_type,
                            "data": encoded
                        }
                    }
                ]
            }
        ],
        "generationConfig": {
            "response_mime_type": "application/json",
            "temperature": 0.1
        }
    }

    url = f"{GEMINI_URL}?key={api_key}"
    headers = {"Content-Type": "application/json"}

    last_error = ""
    gemini_response = None
    successful_model = GEMINI_MODEL

    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=60)
        if resp.status_code == 200:
            gemini_response = resp.json()
        else:
            last_error = f"API returned HTTP {resp.status_code}: {resp.text}"
    except Exception as e:
        last_error = str(e)

    if not gemini_response:
        raise HTTPException(status_code=500, detail=f"Gemini OCR failed: {last_error}")

    try:
        candidates = gemini_response.get("candidates", [])
        if not candidates:
            raise HTTPException(status_code=500, detail="Gemini returned no candidates")

        raw_text = candidates[0]["content"]["parts"][0]["text"]
        parsed = json.loads(raw_text)

        return {
            "semesterStart": parsed.get("semesterStart") or "",
            "semesterEnd": parsed.get("semesterEnd") or "",
            "holidays": parsed.get("holidays") or [],
            "midExams": parsed.get("midExams") or [],
            "labExams": parsed.get("labExams") or [],
            "semesterBreak": parsed.get("semesterBreak") or [],
            "examDates": parsed.get("examDates") or [],
            "studyHolidays": parsed.get("studyHolidays") or [],
            "workingSaturdays": parsed.get("workingSaturdays") or [],
            "events": parsed.get("events") or [],
            "model": successful_model
        }

    except HTTPException:
        raise
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse Gemini JSON response: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gemini calendar parsing failed: {str(e)}")

# --- Additional Features Endpoints ---

@app.post("/auth/google-login", response_model=schemas.Token, tags=["Authentication"])
def google_login(body: dict, db: Session = Depends(get_db)):
    email = body.get("email")
    name = body.get("name")
    if not email or not name:
        raise HTTPException(status_code=400, detail="Invalid google token payload")
    
    user = db.query(models.User).filter(models.User.email == email).first()
    if not user:
        user = models.User(
            name=name,
            email=email,
            password_hash="",  # Empty password for Google OAuth
            college="Google University",
            branch="Computer Science",
            semester="Semester 1 (Autumn)",
            attendance_goal=75.0
        )
        try:
            db.add(user)
            db.commit()
            db.refresh(user)
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to create user profile: {str(e)}")
    
    access_token_expires = auth.timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth.create_access_token(
        data={"sub": user.email}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@app.get("/reports/excel", tags=["Reports"])
def export_report_excel(
    period: str = "monthly",
    start_date: str = None,
    end_date: str = None,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    # Fetch report data (reuse local summary logic)
    report = get_report_summary(period, start_date, end_date, current_user, db)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=10, bold=True)
    regular_font = Font(name="Arial", size=10)
    
    purple_fill = PatternFill(start_color="7C4DFF", end_color="7C4DFF", fill_type="solid")
    dark_gray_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Title Row
    ws.merge_cells("A1:F2")
    title_cell = ws["A1"]
    title_cell.value = "AttendWise Attendance Report"
    title_cell.font = title_font
    title_cell.fill = purple_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Info block
    ws["A4"] = "Student Name:"
    ws["A4"].font = bold_font
    ws["B4"] = report["student"]["name"]
    ws["B4"].font = regular_font
    
    ws["D4"] = "College:"
    ws["D4"].font = bold_font
    ws["E4"] = report["student"]["college"] or "N/A"
    ws["E4"].font = regular_font
    
    ws["A5"] = "Branch/Branch:"
    ws["A5"].font = bold_font
    ws["B5"] = report["student"]["branch"] or "N/A"
    ws["B5"].font = regular_font
    
    ws["D5"] = "Semester:"
    ws["D5"].font = bold_font
    ws["E5"] = report["student"]["semester"] or "N/A"
    ws["E5"].font = regular_font
    
    ws["A6"] = "Period:"
    ws["A6"].font = bold_font
    ws["B6"] = f"{report['start_date']} to {report['end_date']}"
    ws["B6"].font = regular_font
    
    ws["D6"] = "Target Goal:"
    ws["D6"].font = bold_font
    ws["E6"] = f"{report['student']['target_goal']}%"
    ws["E6"].font = regular_font
    
    # Table Header
    headers = ["Subject Name", "Subject Code", "Classes Attended", "Classes Missed", "Total Conducted", "Attendance %"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = dark_gray_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Table Data
    start_row = 9
    for idx, sub in enumerate(report["subjects"]):
        row_idx = start_row + idx
        ws.cell(row=row_idx, column=1, value=sub["name"]).font = regular_font
        ws.cell(row=row_idx, column=2, value=sub["code"] or "N/A").font = regular_font
        ws.cell(row=row_idx, column=3, value=sub["present"]).font = regular_font
        ws.cell(row=row_idx, column=4, value=sub["absent"]).font = regular_font
        ws.cell(row=row_idx, column=5, value=sub["total"]).font = regular_font
        ws.cell(row=row_idx, column=6, value=f"{sub['percentage']}%").font = regular_font
        
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = light_gray_fill
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="right")
    
    # Overall summary row
    summary_row = start_row + len(report["subjects"]) + 1
    ws.cell(row=summary_row, column=1, value="OVERALL SUMMARY").font = bold_font
    ws.cell(row=summary_row, column=3, value=report["overall"]["present"]).font = bold_font
    ws.cell(row=summary_row, column=4, value=report["overall"]["absent"]).font = bold_font
    ws.cell(row=summary_row, column=5, value=report["overall"]["total_conducted"]).font = bold_font
    ws.cell(row=summary_row, column=6, value=f"{report['overall']['percentage']}%").font = bold_font
    
    for col_idx in range(1, 7):
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.border = thin_border
        cell.fill = light_gray_fill
        if col_idx >= 3:
            cell.alignment = Alignment(horizontal="right")
            
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"AttendWise_Report_{period}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/analytics/ai-insights", tags=["Analytics"])
def get_ai_insights(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    # 1. Fetch current subject stats
    subjects = db.query(models.Subject).filter(models.Subject.user_id == current_user.id).all()
    attendances = db.query(models.Attendance).filter(models.Attendance.user_id == current_user.id).all()
    
    sub_summary = []
    for sub in subjects:
        sub_atts = [a for a in attendances if a.subject_id == sub.id]
        p = sum(1 for a in sub_atts if a.status.lower() in ("present", "late_entry", "od", "event_leave", "medical_leave"))
        ab = sum(1 for a in sub_atts if a.status.lower() == "absent")
        total = p + ab
        pct = round(p / total * 100, 1) if total > 0 else 0.0
        sub_summary.append(f"- {sub.name} ({sub.code or ''}): {p} Present, {ab} Absent ({pct}%)")
        
    sub_summary_str = "\n".join(sub_summary)
    
    prompt = f"""You are AttendWise AI, a student attendance companion.
The student "{current_user.name}" has the following attendance record:
{sub_summary_str}

Overall Attendance Goal: {current_user.attendance_goal}%

Analyze their attendance and return a JSON object (no explanations, no markdown block) with these exact fields:
{{
  "recommendations": ["A list of 3-4 specific smart recommendations, e.g. 'Attend next 4 Operating Systems classes', 'You can safely miss 2 Mathematics lectures.'"],
  "risk_analysis": "A brief analysis of their risk of falling below target.",
  "study_suggestions": ["A list of 2-3 suggestions to balance study time with attendance requirements."],
  "general_insight": "A short summary sentence like 'You are safe until next Tuesday.'"
}}
"""
    
    api_key = os.getenv("GEMINI_API_KEY", GEMINI_API_KEY)
    if not api_key:
        return {
            "recommendations": [
                "Attend next 3 class periods to keep your streak active.",
                "Ensure your lab attendance is above 75% before mid-terms."
            ],
            "risk_analysis": "Your attendance is stable, but keep monitoring low-percentage subjects.",
            "study_suggestions": [
                "Utilize weekends to catch up on labs you bunked.",
                "Review class notes for classes marked absent."
            ],
            "general_insight": "You are currently on track to hit your overall goal."
        }
        
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "response_mime_type": "application/json",
            "temperature": 0.2
        }
    }
    
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={api_key}"
    headers = {"Content-Type": "application/json"}
    
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=15)
        if resp.status_code == 200:
            gemini_response = resp.json()
            candidates = gemini_response.get("candidates", [])
            if candidates:
                raw_text = candidates[0]["content"]["parts"][0]["text"]
                return json.loads(raw_text)
    except Exception as e:
        print("Gemini insights generation failed:", e)
        
    return {
        "recommendations": [
            "Attend next 3 class periods to keep your streak active.",
            "Ensure your lab attendance is above 75% before mid-terms."
        ],
        "risk_analysis": "Your attendance is stable, but keep monitoring low-percentage subjects.",
        "study_suggestions": [
            "Utilize weekends to catch up on labs you bunked.",
            "Review class notes for classes marked absent."
        ],
        "general_insight": "You are currently on track to hit your overall goal."
    }

@app.post("/state/restore", tags=["App"])
def restore_state(body: dict, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    user_id = current_user.id
    
    subjects = body.get("subjects", [])
    timetable = body.get("timetable", [])
    attendance_logs = body.get("attendanceLogs", {})
    profile = body.get("profile", {})
    
    if profile:
        current_user.name = profile.get("name", current_user.name)
        current_user.attendance_goal = profile.get("targetGoal", current_user.attendance_goal)
        current_user.semester = profile.get("term", current_user.semester)
        current_user.college = profile.get("college", current_user.college)
        current_user.branch = profile.get("branch", current_user.branch)
        current_user.roll_number = profile.get("roll_number", current_user.roll_number)
        current_user.section = profile.get("section", current_user.section)
        current_user.year = profile.get("year", current_user.year)
        current_user.register_number = profile.get("register_number", current_user.register_number)
        current_user.university = profile.get("university", current_user.university)
        
    db.query(models.Attendance).filter(models.Attendance.user_id == user_id).delete()
    db.query(models.Timetable).filter(models.Timetable.user_id == user_id).delete()
    db.query(models.Subject).filter(models.Subject.user_id == user_id).delete()
    db.commit()
    
    sub_name_to_id = {}
    for sub in subjects:
        db_sub = models.Subject(
            user_id=user_id,
            name=sub["name"],
            code=sub.get("code"),
            prof=sub.get("prof"),
            credits=sub.get("credits", 3),
            color=sub.get("color", "#7c4dff"),
            minimum_required_attendance=sub.get("minimum_required_attendance", 75.0),
            subject_type="Practical" if "lab" in sub["name"].lower() or "practical" in sub["name"].lower() else sub.get("subject_type", "Theory"),
            weekly_classes=sub.get("weekly_classes", 4),
            total_planned_classes=sub.get("total_planned_classes", 40)
        )
        db.add(db_sub)
        db.commit()
        db.refresh(db_sub)
        sub_name_to_id[sub["name"]] = db_sub.id
        
    for tt in timetable:
        sub_name = tt["subject"]
        sub_id = sub_name_to_id.get(sub_name)
        if not sub_id:
            db_sub = models.Subject(user_id=user_id, name=sub_name, color=tt.get("color", "#7c4dff"), subject_type="Practical" if "lab" in sub_name.lower() or "practical" in sub_name.lower() else "Theory")
            db.add(db_sub)
            db.commit()
            db.refresh(db_sub)
            sub_name_to_id[sub_name] = db_sub.id
            sub_id = db_sub.id
            
        start_time_obj = time.fromisoformat(tt["start"])
        end_time_obj = time.fromisoformat(tt["end"])
        
        db_tt = models.Timetable(
            user_id=user_id,
            subject_id=sub_id,
            day=tt["day"],
            start_time=start_time_obj,
            end_time=end_time_obj,
            room=tt.get("room"),
            type=tt.get("type", "Lecture")
        )
        db.add(db_tt)
    db.commit()
    
    for date_str, logs in attendance_logs.items():
        date_obj = date.fromisoformat(date_str)
        for log in logs:
            sub_name = log["subject"]
            sub_id = sub_name_to_id.get(sub_name)
            if not sub_id:
                continue
            
            start_time_obj = time.fromisoformat(log["start"]) if log.get("start") else None
            
            db_att = models.Attendance(
                user_id=user_id,
                subject_id=sub_id,
                date=date_obj,
                start_time=start_time_obj,
                status=log["status"],
                remarks=log.get("remarks")
            )
            db.add(db_att)
            
    db.commit()
    return {"message": "State restored successfully"}

@app.delete("/user/profile", tags=["User"])
def delete_user_profile(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    db.delete(current_user)
    db.commit()
    return {"message": "Account deleted successfully"}

# Android TWA Digital Asset Links Verification
@app.get("/.well-known/assetlinks.json", tags=["Root"])
def get_assetlinks():
    return [{
        "relation": ["delegate_permission/common.handle_all_urls"],
        "target": {
            "namespace": "android_app",
            "package_name": "com.attendwise.app",
            "sha256_cert_fingerprints": [
                "24:7A:25:8E:03:D6:49:3D:AC:F2:FE:A1:FB:05:C6:1E:FA:6D:04:F6:B2:84:25:DE:8C:90:52:5D:65:66:AA:E1"
            ]
        }
    }]

@app.get("/", tags=["Root"])
def read_root():
    return {"message": "Welcome to AttendWise API. Backend is running. Visit /docs for Swagger UI."}


# --- FCM Push Notifications ---
def send_push_notification(fcm_token: str, title: str, body: str):
    if not fcm_token: return
    try:
        from firebase_admin import messaging
        message = messaging.Message(
            notification=messaging.Notification(title=title, body=body),
            token=fcm_token
        )
        messaging.send(message)
    except Exception as e:
        print(f'Failed to send FCM: {e}', flush=True)

import asyncio
from datetime import datetime, timedelta
from .database import SessionLocal

async def notification_scheduler():
    while True:
        try:
            db = SessionLocal()
            now = datetime.now()
            current_day = now.strftime("%a")
            
            start_window = (now + timedelta(minutes=9)).time()
            end_window = (now + timedelta(minutes=11)).time()
            
            upcoming_classes = db.query(models.Timetable).join(models.User).filter(
                models.Timetable.day == current_day,
                models.Timetable.start_time >= start_window,
                models.Timetable.start_time <= end_window
            ).all()
            
            for entry in upcoming_classes:
                print(f"Found upcoming class: {entry.subject.name} for user {entry.owner.email}", flush=True)
                if entry.owner.fcm_token and entry.subject:
                    print(f"Triggering push to {entry.owner.fcm_token}", flush=True)
                    send_push_notification(
                        fcm_token=entry.owner.fcm_token,
                        title="Upcoming Class",
                        body=f"{entry.subject.name} starts in 10 minutes ({entry.start_time.strftime('%H:%M')}). Room: {entry.room or 'TBA'}."
                    )
            
            db.close()
        except Exception as e:
            print(f"Scheduler error: {e}", flush=True)
            
        await asyncio.sleep(30)

@app.on_event("startup")
async def startup_event():
    print("Starting background scheduler...", flush=True)
    asyncio.create_task(notification_scheduler())
